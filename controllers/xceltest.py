import json
from py4web import action, request, response, URL
from py4web.core import redirect
from ..common import db, session, T, flash

import pandas as pd


@action("xceltest/index")
@action.uses("xceltest/index.html")
def index():
    return dict()

from openpyxl import load_workbook

@action("xceltest/upload", method=["POST"])
def upload():
    if "file" not in request.files:
        return json.dumps({"error": "No file uploaded"}), 400

    file = request.files["file"]

    try:
        wb = load_workbook(filename=file.file, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return json.dumps({"error": "Empty file"}), 400

        headers = rows[0]
        data = [dict(zip(headers, row)) for row in rows[1:]]
        return json.dumps(data, default=str)
    except Exception as e:
        return json.dumps({"error": str(e)}), 500
